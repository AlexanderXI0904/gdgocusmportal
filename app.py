import json
import os
import requests
import msal
import atexit
import threading
import uuid
import time
import base64
from pathlib import Path
from flask import Flask, flash, redirect, render_template, request, session, url_for, jsonify
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

# ============================================================
# CONFIGURATION & STATE
# ============================================================
BASE_DIR = Path(__file__).resolve().parent
UPLOAD_DIR = BASE_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
CONFIG_DIR = BASE_DIR / "config"
CONFIG_DIR.mkdir(exist_ok=True)
SETTINGS_FILE = CONFIG_DIR / "settings.json"
CACHE_FILE = CONFIG_DIR / "msal_cache.json"

app = Flask(__name__, template_folder=str(BASE_DIR / "templates"))
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "gdgoc28usm08email26sender2501key")

active_tasks = {}

# ============================================================
# AUTOMATIC CLEANUP
# ============================================================
def cleanup_uploads():
    if UPLOAD_DIR.exists():
        for file in UPLOAD_DIR.glob("*"):
            try:
                if file.is_file():
                    file.unlink()
            except Exception:
                pass

atexit.register(cleanup_uploads)
cleanup_uploads()

DEFAULT_SETTINGS = {
    "client_id": "",
    "tenant_id": "",
    "cc_emails": [],
}

# ============================================================
# SETTINGS & CACHE
# ============================================================
def load_settings():
    if not SETTINGS_FILE.exists():
        save_settings(DEFAULT_SETTINGS.copy())
    try:
        with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
            settings = json.load(f)
            settings.setdefault("client_id", "")
            settings.setdefault("tenant_id", "")
            settings.setdefault("cc_emails", [])
            return settings
    except Exception:
        return DEFAULT_SETTINGS.copy()

def save_settings(settings):
    with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
        json.dump(settings, f, indent=4)

def load_cache():
    cache = msal.SerializableTokenCache()
    if CACHE_FILE.exists():
        try:
            with open(CACHE_FILE, "r") as f:
                cache.deserialize(f.read())
        except Exception:
            pass
    return cache

def save_cache(cache):
    if cache.has_state_changed:
        with open(CACHE_FILE, "w") as f:
            f.write(cache.serialize())

# ============================================================
# MICROSOFT AUTH
# ============================================================
def get_msal_app(cache=None):
    settings = load_settings()
    client_id = settings.get("client_id")
    tenant_id = settings.get("tenant_id")
    if not client_id or not tenant_id:
        return None
    authority = f"https://login.microsoftonline.com/{tenant_id}"
    return msal.PublicClientApplication(client_id, authority=authority, token_cache=cache)

def get_access_token():
    cache = load_cache()
    app = get_msal_app(cache)
    if not app:
        return None
    accounts = app.get_accounts()
    if accounts:
        result = app.acquire_token_silent(["https://graph.microsoft.com/Mail.Send"], account=accounts[0])
        if result and "access_token" in result:
            save_cache(cache)
            return result["access_token"]
    return None

# ============================================================
# EXCEL HELPERS
# ============================================================
def get_workbook():
    filename = session.get("excel_filename")
    # Add strict string validation
    if not isinstance(filename, str) or not filename.strip():
        return None
        
    filepath = UPLOAD_DIR / filename
    if not filepath.exists():
        session.pop("excel_filename", None)
        session.pop("sheet_name", None)
        return None
        
    return load_workbook(filepath, data_only=True)

def get_columns(sheet_name=None):
    wb = get_workbook()
    if not wb: 
        return []
    sheet_name = sheet_name or session.get("sheet_name")
    if not sheet_name or sheet_name not in wb.sheetnames: 
        return []
    ws = wb[sheet_name]
    columns = [str(cell.value).strip() for cell in ws[1] if cell.value is not None and str(cell.value).strip()]
    wb.close()
    return columns

def get_row_data(ws, row_number):
    data = {}
    for index, cell in enumerate(ws[row_number], start=1):
        header_cell = ws.cell(row=1, column=index)
        if header_cell.value is not None:
            col_name = str(header_cell.value).strip()
            if col_name:
                data[col_name] = cell.value
    return data

def process_dynamic_content(text, row_data):
    if not text: 
        return text
    for column, value in row_data.items():
        placeholder = f"{{{{{column}}}}}"
        replacement = "" if value is None else str(value)
        text = text.replace(placeholder, replacement)
    return text

# ============================================================
# BACKGROUND WORKER
# ============================================================
def background_worker(task_id, access_token, subject_template, importance, html_template, email_col, cc_emails, bcc_emails, attachments, sheet_name, excel_filename):
    # Directly load workbook instead of relying on session-based helper
    if not isinstance(excel_filename, str) or not excel_filename.strip():
        active_tasks[task_id]["status"] = "error"
        return
        
    filepath = UPLOAD_DIR / excel_filename
    if not filepath.exists():
        active_tasks[task_id]["status"] = "error"
        return
        
    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception:
        active_tasks[task_id]["status"] = "error"
        return
        
    if sheet_name not in wb.sheetnames:
        active_tasks[task_id]["status"] = "error"
        wb.close()
        return
        
    ws = wb[sheet_name]
    headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}
    
    if email_col not in headers:
        active_tasks[task_id]["status"] = "error"
        wb.close()
        return
        
    email_idx = headers[email_col]
    
    graph_url = "https://graph.microsoft.com/v1.0/me/sendMail"
    headers_http = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"
    }

    # Pre-process attachments to Base64
    b64_attachments = []
    for filename in attachments:
        att_filepath = UPLOAD_DIR / filename
        if att_filepath.exists():
            with open(att_filepath, "rb") as f:
                b64_content = base64.b64encode(f.read()).decode("utf-8")
                b64_attachments.append({
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": filename,
                    "contentBytes": b64_content
                })

    for row_number in range(2, ws.max_row + 1):
        email_val = ws.cell(row=row_number, column=email_idx).value
        if not email_val: continue
        email = str(email_val).strip()
        if not email: continue

        row_data = get_row_data(ws, row_number)
        subject = process_dynamic_content(subject_template, row_data).replace("\n", " ").strip()
        raw_html = process_dynamic_content(html_template, row_data)
        
        message = {
            "message": {
                "subject": subject,
                "importance": importance,
                "body": {"contentType": "HTML", "content": raw_html},
                "toRecipients": [{"emailAddress": {"address": email}}]
            },
            "saveToSentItems": True
        }

        if b64_attachments:
            message["message"]["attachments"] = b64_attachments

        if cc_emails:
            message["message"]["ccRecipients"] = [{"emailAddress": {"address": e.strip()}} for e in cc_emails if e.strip()]
            
        if bcc_emails:
            message["message"]["bccRecipients"] = [{"emailAddress": {"address": e.strip()}} for e in bcc_emails if e.strip()]

        success = False
        for attempt in range(3):
            try:
                resp = requests.post(graph_url, headers=headers_http, json=message, timeout=30)
                if resp.status_code == 202:
                    active_tasks[task_id]["sent"] += 1
                    active_tasks[task_id]["results"].append({"email": email, "status": "Sent"})
                    time.sleep(0.2)
                    success = True
                    break
                elif resp.status_code == 429:
                    retry_after = int(resp.headers.get("Retry-After", 2))
                    time.sleep(retry_after + 1)
                else:
                    break
            except Exception:
                break
                
        if not success:
            active_tasks[task_id]["failed"] += 1
            active_tasks[task_id]["results"].append({"email": email, "status": "Failed"})

    wb.close()
    
    for filename in attachments:
        att_filepath = UPLOAD_DIR / filename
        try:
            if att_filepath.exists():
                att_filepath.unlink()
        except OSError:
            pass

    active_tasks[task_id]["status"] = "completed"

# ============================================================
# ROUTES
# ============================================================
@app.route("/")
def index():
    settings = load_settings()
    authenticated = bool(get_access_token())
    sheets = []
    columns = []

    filename = session.get("excel_filename")
    # Add strict string validation to the condition
    if isinstance(filename, str) and not (UPLOAD_DIR / filename).exists():
        session.pop("excel_filename", None)
        session.pop("sheet_name", None)
        filename = None

    wb = get_workbook()
    if wb:
        sheets = wb.sheetnames
        selected_sheet = session.get("sheet_name")
        if selected_sheet and selected_sheet in sheets:
            columns = [str(c.value).strip() for c in wb[selected_sheet][1] if c.value is not None]
        wb.close()

    return render_template(
        "index.html", 
        page=1, 
        settings=settings, 
        authenticated=authenticated, 
        sheets=sheets, 
        selected_sheet=session.get("sheet_name"), 
        columns=columns, 
        excel_filename=filename if isinstance(filename, str) else None
    )

@app.route("/save-ms-settings", methods=["POST"])
def save_ms_settings():
    settings = load_settings()
    settings["client_id"] = request.form.get("client_id", "").strip()
    settings["tenant_id"] = request.form.get("tenant_id", "").strip()
    save_settings(settings)
    if CACHE_FILE.exists(): 
        CACHE_FILE.unlink()
    flash("Microsoft settings saved.", "success")
    return redirect(url_for("index"))

@app.route("/authenticate")
def authenticate():
    msal_app = get_msal_app()
    if not msal_app:
        flash("Please save your Client ID and Tenant ID first.", "error")
        return redirect(url_for("index"))
        
    flow = msal_app.initiate_device_flow(scopes=["https://graph.microsoft.com/Mail.Send"])
    session["device_flow"] = flow
    return render_template("device_login.html", flow=flow)

@app.route("/complete-authentication", methods=["POST"])
def complete_authentication():
    cache = load_cache()
    msal_app = get_msal_app(cache)
    flow = session.get("device_flow")
    
    if not msal_app or not flow:
        flash("Authentication configuration or flow state missing.", "error")
        return redirect(url_for("index"))
        
    result = msal_app.acquire_token_by_device_flow(flow)
    
    if "access_token" in result:
        save_cache(cache)
        session.pop("device_flow", None)
        flash("Authentication successful.", "success")
    else:
        flash("Authentication failed.", "error")
    return redirect(url_for("index"))

@app.route("/logout")
def logout():
    if CACHE_FILE.exists(): 
        CACHE_FILE.unlink()
    flash("Signed out successfully.", "success")
    return redirect(url_for("index"))

@app.route("/upload-excel", methods=["POST"])
def upload_excel():
    file = request.files.get("file")
    
    if file and file.filename:
        for old_file in UPLOAD_DIR.glob("*"):
            if old_file.suffix.lower() in {".xlsx", ".xlsm"}:
                try: 
                    old_file.unlink()
                except OSError: 
                    pass

        filename = secure_filename(file.filename)
        file.save(UPLOAD_DIR / filename)
        
        session["excel_filename"] = filename
        session.pop("sheet_name", None)
        session.pop("attachments", None)
        flash(f"Excel file '{filename}' uploaded.", "success")
        
    return redirect(url_for("index"))

@app.route("/select-sheet", methods=["POST"])
def select_sheet():
    session["sheet_name"] = request.form.get("sheet_name", "").strip()
    return redirect(url_for("index"))

@app.route("/save-cc", methods=["POST"])
def save_cc():
    email = request.form.get("email", "").strip().lower()
    settings = load_settings()
    cc_emails = settings.setdefault("cc_emails", [])
    if email and email not in cc_emails:
        cc_emails.append(email)
        save_settings(settings)
    return {"success": True}

@app.route("/api/remove-attachment", methods=["POST"])
def api_remove_attachment():
    filename = request.form.get("filename")
    # Add strict string validation
    if not isinstance(filename, str) or not filename.strip():
        return {"success": False}
        
    atts = session.get("attachments", [])
    if filename in atts:
        atts.remove(filename)
        session["attachments"] = atts
        try:
            (UPLOAD_DIR / filename).unlink()
        except OSError:
            pass
    return {"success": True}

@app.route("/editor")
def editor():
    if not get_access_token(): 
        return redirect(url_for("index"))
    return render_template(
        "index.html", 
        page=2, 
        columns=get_columns(), 
        cc_emails=load_settings().get("cc_emails", []), 
        sheet_name=session.get("sheet_name"),
        attachments=session.get("attachments", [])
    )

@app.route("/preview", methods=["POST"])
def preview():
    subject = request.form.get("subject", "").strip()
    importance = request.form.get("importance", "normal")
    html = request.form.get("html", "")
    email_col = request.form.get("email_column", "").strip()
    cc_emails = request.form.getlist("cc_emails")
    bcc_emails = request.form.getlist("bcc_emails")
    
    current_attachments = session.get("attachments", [])
    if 'attachments' in request.files:
        for f in request.files.getlist('attachments'):
            if f.filename:
                safe_name = secure_filename(f.filename)
                f.save(UPLOAD_DIR / safe_name)
                if safe_name not in current_attachments:
                    current_attachments.append(safe_name)
    
    session["attachments"] = current_attachments
    saved_attachments = current_attachments

    wb = get_workbook()
    sheet_name = session.get("sheet_name")
    
    if not wb or not isinstance(sheet_name, str) or sheet_name not in wb.sheetnames:
        if wb:
            wb.close()
        flash("System error: Worksheet data stream is unavailable.", "error")
        return redirect(url_for("index"))

    ws = wb[sheet_name]
    headers = {str(c.value).strip(): idx for idx, c in enumerate(ws[1], start=1) if c.value is not None}
    
    if email_col not in headers:
        wb.close()
        flash("System error: Selected mapping column no longer exists.", "error")
        return redirect(url_for("editor"))
        
    email_idx = headers[email_col]
    
    preview_data = []
    recipients = []
    for row_num in range(2, ws.max_row + 1):
        email_val = ws.cell(row=row_num, column=email_idx).value
        if email_val:
            email = str(email_val).strip()
            if email:
                row_data = get_row_data(ws, row_num)
                recipients.append(email)
                preview_data.append({"email": email, "data": {k: ("" if v is None else str(v)) for k, v in row_data.items()}})
    
    wb.close()

    return render_template(
        "index.html", 
        page=3, 
        subject=subject, 
        importance=importance, 
        html=html, 
        email_column=email_col, 
        cc_emails=cc_emails, 
        bcc_emails=bcc_emails, 
        attachments=saved_attachments, 
        sheet_name=sheet_name, 
        recipient_count=len(recipients), 
        recipients=recipients, 
        preview_data_json=json.dumps(preview_data)
    )

@app.route("/start-send", methods=["POST"])
def start_send():
    access_token = get_access_token()
    if not access_token:
        flash("Authentication expired.", "error")
        return redirect(url_for("index"))

    # Extract filename on the main thread
    excel_filename = session.get("excel_filename")
    if not isinstance(excel_filename, str) or not excel_filename.strip():
        flash("Excel file data lost. Please restart your campaign.", "error")
        return redirect(url_for("index"))

    task_id = str(uuid.uuid4())
    active_tasks[task_id] = {
        "status": "running", 
        "total": int(request.form.get("total_count", 0)), 
        "sent": 0, 
        "failed": 0, 
        "results": []
    }

    attachments_to_send = request.form.getlist("attachments")
    session.pop("attachments", None)

    kwargs = {
        "task_id": task_id,
        "access_token": access_token,
        "subject_template": request.form.get("subject", ""),
        "importance": request.form.get("importance", "normal"),
        "html_template": request.form.get("html", ""),
        "email_col": request.form.get("email_column", ""),
        "cc_emails": request.form.getlist("cc_emails"),
        "bcc_emails": request.form.getlist("bcc_emails"),
        "attachments": attachments_to_send,
        "sheet_name": session.get("sheet_name"),
        "excel_filename": excel_filename  # Pass extracted string directly
    }

    threading.Thread(target=background_worker, kwargs=kwargs).start()
    
    return render_template("index.html", page=4, task_id=task_id)

@app.route("/task-status/<task_id>")
def task_status(task_id):
    task = active_tasks.get(task_id, {"status": "not_found"})
    return jsonify(task)

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=True)